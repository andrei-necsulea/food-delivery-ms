import csv
from io import BytesIO, StringIO
from django.http import HttpResponse

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None


def export_csv(filename, headers, rows):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)

    response = HttpResponse(buffer.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    return response


def export_xlsx(filename, headers, rows):
    if openpyxl is None:
        raise RuntimeError('openpyxl is required for XLSX exports')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    # auto width
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col:
            try:
                v = str(cell.value)
            except Exception:
                v = ''
            if v and len(v) > max_length:
                max_length = len(v)
        ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    response = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response
